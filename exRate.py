from ast import While
from datetime import date
from bs4 import BeautifulSoup
import urllib.request
import csv
import time
import os
import sys
from funcfile import func
import openpyxl as opx


# 把网址 URL 存在变量里
urlpage = 'https://fx.sauder.ubc.ca/today.html'
filename = 'dailyrate.xlsx'
Todaytime = time.strftime("%Y-%m-%d",time.localtime(time.time()))
ratetype='SP'
# 获取网页内容，把 HTML 数据保存在 page 变量中
page = urllib.request.urlopen(urlpage)

# 用 Beautiful Soup 解析 html 数据，
# 并保存在 soup 变量里
soup = BeautifulSoup(page, 'html.parser')

# 在表格中查找数据
table = soup.find('table', attrs={'class': 'fxt'})
results = table.find_all('tr')
print('Number of results', len(results))


# 创建一个列表对象，并且把表头数据作为列表的第一个元素.并且判断数据表格是否已经生成，如果已经生成，便不会再添加了。

rows = []
path=sys.path[0]
print(path,filename)
cfile=func.CheckFile(path,filename)

if cfile == False:
    rows.append(['Date','Code','Currency','Type','fcu/CAD','CAD/fcu',
            'fcu/USD','USD/fcu','fcu/EUR','EUR/fcu']
             )

# 遍历所有数据
for result in results:
# 找到每一个 td 单元格的内容
    data = result.find_all('td')
# 如果该单元格无数据，则跳过
    if len(data) == 0: 
        continue

    # 接上图，将单元格内容保存到变量中
    CurrencyCode = data[0].getText()
    CurrencyName = data[1].getText()
    fcucad = data[2].getText()
    cadfcu = data[3].getText()
    fcuusd = data[4].getText()
    usdfcu = data[5].getText()
    fcueur = data[6].getText()
    eurfcu = data[7].getText()
    if CurrencyCode == 'USD':
            rows.append([Todaytime,CurrencyCode,CurrencyName,ratetype,fcucad,cadfcu,fcuusd,usdfcu,fcueur,eurfcu])
    elif CurrencyCode == 'CAD':
        rows.append([Todaytime,CurrencyCode,CurrencyName,ratetype,fcucad,cadfcu,fcuusd,usdfcu,fcueur,eurfcu])

print(rows)
datafile = path+r"\\"+filename
if cfile == False:
    wb = opx.Workbook()
    wb = opx.Workbook(datafile)
    ws1 = wb.create_sheet("Sheet1", 0)
else:
    wb = opx.load_workbook(datafile)
    ws1 = wb.active
    ws1 = wb['Sheet1']
#判断，最后的时间记录，如果已经存在，便不进行更新。
maxrows=ws1.max_row
res1=ws1.cell(row=maxrows,column=1).value
if res1 == Todaytime:
    print('You already downloaded todays exchange rate, the sofaware will closed')
    
else:        
    for row in rows:
            ws1.append(row)
    wb.save(datafile)
